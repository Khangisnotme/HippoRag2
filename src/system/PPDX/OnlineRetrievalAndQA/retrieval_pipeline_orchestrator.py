"""
OnlineRetrievalAndQA/retrieval_pipeline_orchestrator.py
Retrieval Pipeline Orchestrator - Điều phối toàn bộ quá trình retrieval
Kết hợp Module 1, 2, 3, 4 để tạo ra complete retrieval pipeline

Mô tả chi tiết:
    Orchestrator này điều phối việc thực thi các modules theo thứ tự:
    Module 1 (Dual Retrieval) → Module 2 (Triple Filtering) → 
    Module 3 (Passage Ranking) → Module 4 (Context Expansion - Optional)
    
Workflow:
    Query → Dual Retrieval → Triple Filtering → Passage Ranking → [Context Expansion] → Final Result

Output Format cho So sánh Baseline:
    Trả ra standardized format để so sánh với baseline RAG:
    {
        "query_id": "Q001",
        "query": "Lợi ích của táo cho sức khỏe",
        "top_passages": [
            {"passage_id": "P001", "rank": 1, "score": 0.85, "text": "..."},
            {"passage_id": "P015", "rank": 2, "score": 0.82, "text": "..."},
            {"passage_id": "P007", "rank": 3, "score": 0.78, "text": "..."},
            # ... đến rank 10
        ],
        "method": "proposed_hipporag",
        "processing_time": 12.5,
        "statistics": {...}
    }

Evaluation Metrics:
    So sánh với baseline RAG sử dụng:
    - Precision@K: Trong top-K passages, có bao nhiêu passages relevant?
    - Recall@K: Có ít nhất 1 relevant passage trong top-K không?
    - MRR: Mean Reciprocal Rank của relevant passage đầu tiên
    
    Ví dụ so sánh:
    Baseline RAG output:     ["P015", "P032", "P001", "P088", "P007", ...]
    Proposed method output:  ["P001", "P007", "P015", "P025", "P044", ...]
    
    Với ground truth: ["P001", "P007", "P025"]
    Precision@3: Baseline=1/3=0.33, Proposed=3/3=1.00 (+200% improvement)
    Recall@5: Baseline=2/3=0.67, Proposed=3/3=1.00 (+49% improvement)
"""

from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import logging
import time
import json
from dataclasses import dataclass, asdict
from datetime import datetime
import traceback
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Module imports
from module1_dual_retrieval import DualRetriever, RetrievalConfig, RetrievalResult
from module2_triple_filter import LLMTripleFilter, TripleFilterConfig, FilteringResult
from module3_passage_ranker import PassageRanker, PassageRankerConfig, RankingResult
from module4_context_expander import ContextExpander, ContextExpansionConfig, ExpansionResult

# Shared utilities
from utils.utils_shared_general import (
    setup_logger,
    log_performance,
    validate_query,
    save_json,
    load_json,
    create_query_metadata,
    PerformanceStats
)

# Create log directory if it doesn't exist
log_dir = Path("outputs/log")
log_dir.mkdir(parents=True, exist_ok=True)

# Setup logger with file output
log_file = log_dir / f"retrieval_pipeline_orchestrator_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logger = setup_logger(__name__, log_file=log_file)

# ==================== DATA CLASSES ====================

@dataclass
class RetrievalPipelineConfig:
    """
    Cấu hình toàn diện cho Retrieval Pipeline
    
    Attributes:
        # Module configurations
        dual_retrieval_config (RetrievalConfig): Cấu hình Module 1 - Dual Retrieval
        triple_filter_config (TripleFilterConfig): Cấu hình Module 2 - Triple Filtering
        passage_ranking_config (PassageRankerConfig): Cấu hình Module 3 - Passage Ranking
        context_expansion_config (ContextExpansionConfig): Cấu hình Module 4 - Context Expansion
        
        # Pipeline control
        enable_context_expansion (bool): Có chạy Module 4 không
        max_final_passages (int): Số lượng passages tối đa trong kết quả cuối
        enable_intermediate_saving (bool): Có lưu kết quả trung gian không
        
        # Output formatting
        output_format (str): Format output ("standard", "evaluation", "detailed")
        include_metadata (bool): Có include metadata chi tiết không
        include_intermediate_results (bool): Có include kết quả trung gian không
        
        # Performance settings
        timeout_seconds (int): Timeout cho toàn bộ pipeline
        enable_performance_tracking (bool): Có track performance không
    """
    # Module configurations
    dual_retrieval_config: RetrievalConfig = None
    triple_filter_config: TripleFilterConfig = None
    passage_ranking_config: PassageRankerConfig = None
    context_expansion_config: ContextExpansionConfig = None
    
    # Pipeline control
    enable_context_expansion: bool = False
    max_final_passages: int = 10
    enable_intermediate_saving: bool = False
    
    # Output formatting for evaluation
    output_format: str = "evaluation"  # "standard", "evaluation", "detailed"
    include_metadata: bool = True
    include_intermediate_results: bool = False
    
    # Performance settings
    timeout_seconds: int = 300  # 5 minutes timeout
    enable_performance_tracking: bool = True
    
    def __post_init__(self):
        """Initialize default configs nếu không được provide"""
        if self.dual_retrieval_config is None:
            self.dual_retrieval_config = RetrievalConfig()
        if self.triple_filter_config is None:
            self.triple_filter_config = TripleFilterConfig()
        if self.passage_ranking_config is None:
            self.passage_ranking_config = PassageRankerConfig()
        if self.context_expansion_config is None:
            self.context_expansion_config = ContextExpansionConfig()

@dataclass
class RetrievalPipelineResult:
    """
    Kết quả hoàn chỉnh của Retrieval Pipeline
    
    Attributes:
        query_id (str): ID unique của query
        query (str): Query gốc của user
        top_passages (List[Dict]): Top passages được ranked (format chuẩn cho evaluation)
        method (str): Tên method ("proposed_hipporag")
        processing_time (float): Tổng thời gian xử lý
        
        # Intermediate results (optional)
        raw_retrieval_result (RetrievalResult): Kết quả Module 1
        filtering_result (FilteringResult): Kết quả Module 2
        ranking_result (RankingResult): Kết quả Module 3
        expansion_result (ExpansionResult): Kết quả Module 4 (nếu có)
        
        # Statistics và metadata
        statistics (Dict[str, Any]): Thống kê chi tiết
        config_used (RetrievalPipelineConfig): Config được sử dụng
        errors (List[Dict]): Danh sách errors nếu có
    """
    query_id: str
    query: str
    top_passages: List[Dict[str, Any]]
    method: str
    processing_time: float
    
    # Intermediate results
    raw_retrieval_result: Optional[RetrievalResult] = None
    filtering_result: Optional[FilteringResult] = None
    ranking_result: Optional[RankingResult] = None
    expansion_result: Optional[ExpansionResult] = None
    
    # Statistics
    statistics: Dict[str, Any] = None
    config_used: RetrievalPipelineConfig = None
    errors: List[Dict] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert thành dictionary cho serialization"""
        result = {
            'query_id': self.query_id,
            'query': self.query,
            'top_passages': self.top_passages,
            'method': self.method,
            'processing_time': self.processing_time,
            'statistics': self.statistics or {},
            'errors': self.errors or []
        }
        
        # Add intermediate results nếu available
        if self.raw_retrieval_result:
            result['raw_retrieval_result'] = self.raw_retrieval_result.to_dict()
        if self.filtering_result:
            result['filtering_result'] = self.filtering_result.to_dict()
        if self.ranking_result:
            result['ranking_result'] = self.ranking_result.to_dict()
        if self.expansion_result:
            result['expansion_result'] = self.expansion_result.to_dict()
        if self.config_used:
            result['config_used'] = asdict(self.config_used)
        
        return result
    
    def save_to_file(self, filepath: Path):
        """Lưu kết quả vào file JSON"""
        save_json(self.to_dict(), filepath, indent=2)
        logger.info(f"💾 Đã lưu pipeline result vào: {filepath}")
    
    def get_evaluation_format(self) -> Dict[str, Any]:
        """
        Trả về format chuẩn cho evaluation với baseline
        
        Returns:
            Dict: Format chuẩn để so sánh metrics
        """
        return {
            "query_id": self.query_id,
            "query": self.query,
            "method": self.method,
            "top_passages": self.top_passages,
            "processing_time": self.processing_time,
            "metadata": {
                "total_passages_retrieved": len(self.raw_retrieval_result.raw_passages) if self.raw_retrieval_result else 0,
                "filtered_triples_count": len(self.filtering_result.filtered_triples) if self.filtering_result else 0,
                "final_passages_count": len(self.top_passages)
            }
        }

    def to_excel_summary_row(self) -> Dict[str, Any]:
        """
        Convert result thành row cho Excel summary sheet
        
        Returns:
            Dict: Row data cho summary sheet
        """
        top_scores = [p['score'] for p in self.top_passages] if self.top_passages else [0]
        
        return {
            'Query ID': self.query_id,
            'Query Text': self.query[:100] + ('...' if len(self.query) > 100 else ''),
            'Method': self.method,
            'Processing Time (s)': round(self.processing_time, 2),
            'Top Passages Count': len(self.top_passages),
            'Top Score': round(max(top_scores), 3) if top_scores else 0,
            'Average Score': round(sum(top_scores) / len(top_scores), 3) if top_scores else 0,
            'Min Score': round(min(top_scores), 3) if top_scores else 0,
            'Score Range': round(max(top_scores) - min(top_scores), 3) if len(top_scores) > 1 else 0
        }

    def to_excel_detailed_rows(self) -> List[Dict[str, Any]]:
        """
        Convert result thành rows cho Excel detailed sheet
        
        Returns:
            List[Dict]: Rows data cho detailed sheet
        """
        rows = []
        for passage in self.top_passages:
            metadata = passage.get('metadata', {})
            
            row = {
                'Query ID': self.query_id,
                'Query Text': self.query[:50] + ('...' if len(self.query) > 50 else ''),
                'Rank': passage.get('rank', 0),
                'Passage ID': passage.get('passage_id', ''),
                'Final Score': round(passage.get('score', 0), 3),
                'Retrieval Score': round(metadata.get('hybrid_retrieval_score', 0), 3),
                'Support Score': round(metadata.get('support_score', 0), 3),
                'Supporting Triples': metadata.get('supporting_triples_count', 0),
                'Text Preview': passage.get('text', '')[:150] + ('...' if len(passage.get('text', '')) > 150 else ''),
                'Text Length': len(passage.get('text', '')),
                'Score Breakdown': str(metadata.get('score_breakdown', {}))
            }
            rows.append(row)
        
        return rows

# ==================== PIPELINE ORCHESTRATOR ====================

class RetrievalPipelineOrchestrator:
    """
    Main Orchestrator cho Retrieval Pipeline
    
    Điều phối việc thực thi tuần tự các modules và tạo ra output chuẩn
    để so sánh với baseline RAG systems.
    """
    
    def __init__(self, config: Optional[RetrievalPipelineConfig] = None,
                 neo4j_uri: str = "bolt://localhost:7687",
                 neo4j_user: str = "neo4j", 
                 neo4j_password: str = "graphrag123"):
        """
        Khởi tạo Pipeline Orchestrator
        
        Args:
            config (Optional[RetrievalPipelineConfig]): Cấu hình pipeline
            neo4j_uri (str): URI Neo4j database
            neo4j_user (str): Neo4j username
            neo4j_password (str): Neo4j password
        """
        self.config = config or RetrievalPipelineConfig()
        self.neo4j_uri = neo4j_uri
        self.neo4j_user = neo4j_user
        self.neo4j_password = neo4j_password
        
        # Initialize modules
        self.dual_retriever = None
        self.triple_filter = None
        self.passage_ranker = None
        self.context_expander = None
        
        # Performance tracking
        self.module_times = {}
        self.errors = []
        
        logger.info("🚀 Khởi tạo RetrievalPipelineOrchestrator...")
        logger.info(f"   🔧 Context expansion: {'Bật' if self.config.enable_context_expansion else 'Tắt'}")
        logger.info(f"   📊 Output format: {self.config.output_format}")
        logger.info(f"   📋 Max final passages: {self.config.max_final_passages}")
        
        self._initialize_modules()
    
    def _initialize_modules(self):
        """Khởi tạo tất cả các modules"""
        try:
            logger.info("🔧 Đang khởi tạo các modules...")
            
            # Module 1: Dual Retriever
            logger.info("   📊 Khởi tạo Module 1 - Dual Retrieval...")
            self.dual_retriever = DualRetriever(
                self.neo4j_uri, self.neo4j_user, self.neo4j_password,
                self.config.dual_retrieval_config
            )
            
            # Module 2: Triple Filter
            logger.info("   🤖 Khởi tạo Module 2 - LLM Triple Filter...")
            self.triple_filter = LLMTripleFilter(self.config.triple_filter_config)
            
            # Module 3: Passage Ranker
            logger.info("   🏆 Khởi tạo Module 3 - Passage Ranker...")
            self.passage_ranker = PassageRanker(self.config.passage_ranking_config)
            
            # Module 4: Context Expander (nếu enabled)
            if self.config.enable_context_expansion:
                logger.info("   🔗 Khởi tạo Module 4 - Context Expander...")
                self.context_expander = ContextExpander(self.config.context_expansion_config)
            
            logger.info("✅ Tất cả modules đã được khởi tạo thành công")
            
        except Exception as e:
            logger.error(f"❌ Lỗi khởi tạo modules: {e}")
            self.errors.append({
                'stage': 'module_initialization',
                'error': str(e),
                'timestamp': time.time()
            })
            raise
    
    def process_single_query(self, query: str, query_id: Optional[str] = None) -> RetrievalPipelineResult:
        """
        Xử lý một query duy nhất qua toàn bộ pipeline
        
        Args:
            query (str): Query của user
            query_id (Optional[str]): ID của query (auto-generate nếu None)
            
        Returns:
            RetrievalPipelineResult: Kết quả hoàn chỉnh
        """
        if query_id is None:
            query_id = f"Q_{int(time.time() * 1000)}"
        
        logger.info("=" * 80)
        logger.info("🚀 BẮT ĐẦU RETRIEVAL PIPELINE")
        logger.info("=" * 80)
        logger.info(f"📝 Query ID: {query_id}")
        logger.info(f"📝 Query: '{query}'")
        
        start_time = time.time()
        
        # Validate query
        if not validate_query(query):
            error_msg = f"Invalid query: {query}"
            logger.error(f"❌ {error_msg}")
            return self._create_error_result(query_id, query, error_msg)
        
        try:
            # Module 1: Dual Retrieval
            logger.info(f"\n📊 MODULE 1: DUAL RETRIEVAL")
            logger.info("-" * 60)
            module1_start = time.time()
            
            retrieval_result = self.dual_retriever.retrieve_dual(
                query,
                top_k_passages=self.config.dual_retrieval_config.max_passages,
                top_n_triples=10  # Retrieve nhiều để filter
            )
            
            module1_time = time.time() - module1_start
            self.module_times['dual_retrieval'] = module1_time
            logger.info(f"✅ Module 1 hoàn thành trong {module1_time:.2f}s")
            
            # Module 2: Triple Filtering
            logger.info(f"\n🤖 MODULE 2: LLM TRIPLE FILTERING")
            logger.info("-" * 60)
            module2_start = time.time()
            
            filtering_result = self.triple_filter.filter_triples(
                query, retrieval_result.raw_triples
            )
            
            module2_time = time.time() - module2_start
            self.module_times['triple_filtering'] = module2_time
            logger.info(f"✅ Module 2 hoàn thành trong {module2_time:.2f}s")
            
            # Module 3: Passage Ranking
            logger.info(f"\n🏆 MODULE 3: PASSAGE RANKING")
            logger.info("-" * 60)
            module3_start = time.time()
            
            ranking_result = self.passage_ranker.rank_passages(
                retrieval_result.raw_passages,
                filtering_result.filtered_triples,
                query
            )
            
            module3_time = time.time() - module3_start
            self.module_times['passage_ranking'] = module3_time
            logger.info(f"✅ Module 3 hoàn thành trong {module3_time:.2f}s")
            
            # Module 4: Context Expansion (Optional)
            expansion_result = None
            if self.config.enable_context_expansion:
                logger.info(f"\n🔗 MODULE 4: CONTEXT EXPANSION")
                logger.info("-" * 60)
                module4_start = time.time()
                
                expansion_result = self.context_expander.expand_context(
                    filtering_result.filtered_triples, query
                )
                
                module4_time = time.time() - module4_start
                self.module_times['context_expansion'] = module4_time
                logger.info(f"✅ Module 4 hoàn thành trong {module4_time:.2f}s")
            
            # Generate final result
            total_time = time.time() - start_time
            
            final_result = self._create_final_result(
                query_id, query, total_time,
                retrieval_result, filtering_result, ranking_result, expansion_result
            )
            
            # Log final summary
            self._log_pipeline_summary(final_result)
            
            return final_result
            
        except Exception as e:
            logger.error(f"❌ Lỗi trong pipeline: {e}")
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            
            error_info = {
                'stage': 'pipeline_execution',
                'error': str(e),
                'traceback': traceback.format_exc(),
                'timestamp': time.time()
            }
            self.errors.append(error_info)
            
            return self._create_error_result(query_id, query, str(e))
    
    def process_multiple_queries(self, queries: List[Dict[str, str]]) -> List[RetrievalPipelineResult]:
        """
        Xử lý multiple queries cho evaluation
        
        Args:
            queries (List[Dict]): List of {"query_id": "Q001", "query": "text"}
            
        Returns:
            List[RetrievalPipelineResult]: Danh sách kết quả
        """
        logger.info(f"🔄 Bắt đầu xử lý {len(queries)} queries...")
        
        results = []
        for i, query_data in enumerate(queries, 1):
            query_id = query_data.get('query_id', f'Q_{i:03d}')
            query_text = query_data.get('query', '')
            
            logger.info(f"\n📊 PROCESSING QUERY {i}/{len(queries)}: {query_id}")
            
            try:
                result = self.process_single_query(query_text, query_id)
                results.append(result)
                
                logger.info(f"✅ Query {query_id} hoàn thành")
                
                # Save intermediate nếu enabled
                if self.config.enable_intermediate_saving:
                    output_path = Path(f"outputs/intermediate/query_{query_id}_result.json")
                    result.save_to_file(output_path)
                
            except Exception as e:
                logger.error(f"❌ Lỗi xử lý query {query_id}: {e}")
                error_result = self._create_error_result(query_id, query_text, str(e))
                results.append(error_result)
        
        logger.info(f"🎉 Hoàn thành xử lý {len(results)} queries")
        return results
    
    def _create_final_result(self, query_id: str, query: str, total_time: float,
                           retrieval_result: RetrievalResult,
                           filtering_result: FilteringResult,
                           ranking_result: RankingResult,
                           expansion_result: Optional[ExpansionResult]) -> RetrievalPipelineResult:
        """Tạo final result object"""
        
        # Format top passages cho evaluation
        top_passages = self._format_passages_for_evaluation(ranking_result.ranked_passages)
        
        # Create statistics
        statistics = self._create_pipeline_statistics(
            retrieval_result, filtering_result, ranking_result, expansion_result, total_time
        )
        
        return RetrievalPipelineResult(
            query_id=query_id,
            query=query,
            top_passages=top_passages,
            method="proposed_hipporag",
            processing_time=total_time,
            raw_retrieval_result=retrieval_result if self.config.include_intermediate_results else None,
            filtering_result=filtering_result if self.config.include_intermediate_results else None,
            ranking_result=ranking_result if self.config.include_intermediate_results else None,
            expansion_result=expansion_result if self.config.include_intermediate_results else None,
            statistics=statistics,
            config_used=self.config,
            errors=self.errors.copy()
        )
    
    def _format_passages_for_evaluation(self, ranked_passages) -> List[Dict[str, Any]]:
        """
        Format passages thành format chuẩn cho evaluation
        
        Args:
            ranked_passages: List of RankedPassage objects
            
        Returns:
            List[Dict]: Standardized format cho so sánh với baseline
        """
        formatted = []
        
        for passage in ranked_passages[:self.config.max_final_passages]:
            passage_dict = {
                "passage_id": passage.passage_id,
                "rank": passage.rank,
                "score": passage.final_score,
                "text": passage.original_text
            }
            
            # Add metadata nếu enabled
            if self.config.include_metadata:
                passage_dict["metadata"] = {
                    "hybrid_retrieval_score": passage.hybrid_retrieval_score,
                    "support_score": passage.support_score,
                    "supporting_triples_count": passage.supporting_triples_count,
                    "score_breakdown": passage.score_breakdown
                }
            
            formatted.append(passage_dict)
        
        return formatted
    
    def _create_pipeline_statistics(self, retrieval_result, filtering_result, 
                                  ranking_result, expansion_result, total_time) -> Dict[str, Any]:
        """Tạo comprehensive statistics"""
        
        stats = {
            'pipeline_overview': {
                'total_time_seconds': total_time,
                'modules_executed': len(self.module_times),
                'context_expansion_enabled': self.config.enable_context_expansion
            },
            'module_times': self.module_times.copy(),
            'data_flow': {
                'raw_passages': len(retrieval_result.raw_passages) if retrieval_result else 0,
                'raw_triples': len(retrieval_result.raw_triples) if retrieval_result else 0,
                'filtered_triples': len(filtering_result.filtered_triples) if filtering_result else 0,
                'final_passages': len(ranking_result.ranked_passages) if ranking_result else 0,
                'expanded_contexts': len(expansion_result.expanded_contexts) if expansion_result else 0
            },
            'efficiency_metrics': {
                'triple_filtering_rate': (
                    len(filtering_result.filtered_triples) / len(retrieval_result.raw_triples)
                    if filtering_result and retrieval_result and len(retrieval_result.raw_triples) > 0
                    else 0
                ),
                'passage_reranking_impact': (
                    len(ranking_result.ranked_passages) / len(retrieval_result.raw_passages)
                    if ranking_result and retrieval_result and len(retrieval_result.raw_passages) > 0
                    else 0
                )
            },
            'errors_count': len(self.errors)
        }
        
        return stats
    
    def _create_error_result(self, query_id: str, query: str, error_msg: str) -> RetrievalPipelineResult:
        """Tạo error result khi có lỗi"""
        return RetrievalPipelineResult(
            query_id=query_id,
            query=query,
            top_passages=[],
            method="proposed_hipporag",
            processing_time=0.0,
            statistics={'error': error_msg},
            errors=[{'error': error_msg, 'timestamp': time.time()}]
        )
    
    def _log_pipeline_summary(self, result: RetrievalPipelineResult):
        """Log tổng kết pipeline execution"""
        logger.info("=" * 80)
        logger.info("🎉 RETRIEVAL PIPELINE HOÀN THÀNH")
        logger.info("=" * 80)
        logger.info(f"📝 Query: '{result.query}'")
        logger.info(f"⏱️ Tổng thời gian: {result.processing_time:.2f}s")
        logger.info(f"📊 Top passages: {len(result.top_passages)}")
        
        # Module times breakdown
        logger.info(f"\n📈 THỜI GIAN TỪNG MODULE:")
        for module, time_taken in self.module_times.items():
            percentage = (time_taken / result.processing_time * 100) if result.processing_time > 0 else 0
            logger.info(f"   {module}: {time_taken:.2f}s ({percentage:.1f}%)")
        
        # Top passages preview
        if result.top_passages:
            logger.info(f"\n🏆 TOP 3 PASSAGES:")
            for i, passage in enumerate(result.top_passages[:3], 1):
                logger.info(f"   {i}. {passage['passage_id']}: {passage['score']:.3f}")
                logger.info(f"      📝 {passage['text'][:100]}...")
        
        # Data flow
        if result.statistics and 'data_flow' in result.statistics:
            flow = result.statistics['data_flow']
            logger.info(f"\n📊 DATA FLOW:")
            logger.info(f"   Raw passages: {flow['raw_passages']}")
            logger.info(f"   Raw triples: {flow['raw_triples']}")
            logger.info(f"   Filtered triples: {flow['filtered_triples']}")
            logger.info(f"   Final passages: {flow['final_passages']}")
            if flow['expanded_contexts'] > 0:
                logger.info(f"   Expanded contexts: {flow['expanded_contexts']}")
        
        logger.info("=" * 80)
    
    def close(self):
        """Đóng tất cả connections và cleanup"""
        logger.info("🔐 Đang đóng Pipeline Orchestrator...")
        
        if self.dual_retriever:
            self.dual_retriever.close()
        if self.context_expander:
            self.context_expander.close()
        
        logger.info("✅ Pipeline Orchestrator đã được đóng")

# ==================== UTILITY FUNCTIONS ====================

def create_default_pipeline_config() -> RetrievalPipelineConfig:
    """Tạo cấu hình mặc định cho pipeline"""
    return RetrievalPipelineConfig(
        enable_context_expansion=False,  # Tắt context expansion để focus vào retrieval
        max_final_passages=10,
        enable_intermediate_saving=False,
        output_format="evaluation",
        include_metadata=True,
        include_intermediate_results=False,
        enable_performance_tracking=True
    )

def load_queries_from_file(filepath: Path) -> List[Dict[str, str]]:
    """
    Load queries từ file JSON
    
    Args:
        filepath (Path): Đường dẫn file chứa queries
        
    Returns:
        List[Dict]: List of {"query_id": "Q001", "query": "text"}
    """
    if not filepath.exists():
        raise FileNotFoundError(f"Query file not found: {filepath}")
    
    data = load_json(filepath)
    
    # Support multiple formats
    if isinstance(data, list):
        # Format: [{"query_id": "Q001", "query": "text"}, ...]
        return data
    elif isinstance(data, dict):
        if 'queries' in data:
            # Format: {"queries": [...]}
            return data['queries']
        else:
            # Format: {"Q001": "text", "Q002": "text", ...}
            return [{"query_id": qid, "query": text} for qid, text in data.items()]
    
    raise ValueError("Invalid query file format")

def save_evaluation_results(results: List[RetrievalPipelineResult], filepath: Path):
    """
    Lưu results theo format evaluation chuẩn
    
    Args:
        results (List[RetrievalPipelineResult]): Pipeline results
        filepath (Path): Output file path
    """
    evaluation_data = {
        "method": "proposed_hipporag",
        "total_queries": len(results),
        "timestamp": datetime.now().isoformat(),
        "results": [result.get_evaluation_format() for result in results]
    }
    
    save_json(evaluation_data, filepath, indent=2)
    logger.info(f"💾 Đã lưu evaluation results vào: {filepath}")
    logger.info(f"   📊 Total queries: {len(results)}")

def save_results_to_excel(results: List[RetrievalPipelineResult], 
                         filepath: Path,
                         include_charts: bool = True) -> None:
    """
    Save pipeline results thành Excel file với multiple sheets và formatting
    
    Args:
        results (List[RetrievalPipelineResult]): Pipeline results
        filepath (Path): Excel output file path
        include_charts (bool): Có tạo charts không
    """
    logger.info(f"📊 Saving {len(results)} results to Excel: {filepath}")
    
    # Ensure output directory exists
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Sheet 1: Summary
        logger.info("   📋 Creating Summary sheet...")
        summary_data = [result.to_excel_summary_row() for result in results]
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Detailed Results
        logger.info("   📄 Creating Detailed Results sheet...")
        detailed_data = []
        for result in results:
            detailed_data.extend(result.to_excel_detailed_rows())
        
        if detailed_data:
            detailed_df = pd.DataFrame(detailed_data)
            detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)
        
        # Sheet 3: Performance Statistics
        logger.info("   📈 Creating Performance Stats sheet...")
        perf_stats = _calculate_performance_statistics(results)
        perf_df = pd.DataFrame(list(perf_stats.items()), columns=['Metric', 'Value'])
        perf_df.to_excel(writer, sheet_name='Performance Stats', index=False)
        
        # Sheet 4: Module Breakdown
        logger.info("   🔧 Creating Module Breakdown sheet...")
        module_data = _create_module_breakdown_data(results)
        if module_data:
            module_df = pd.DataFrame(module_data)
            module_df.to_excel(writer, sheet_name='Module Breakdown', index=False)
        
        # Sheet 5: Score Distribution
        logger.info("   📊 Creating Score Distribution sheet...")
        score_data = _create_score_distribution_data(results)
        if score_data:
            score_df = pd.DataFrame(score_data)
            score_df.to_excel(writer, sheet_name='Score Distribution', index=False)
    
    # Apply formatting
    _format_excel_workbook(filepath, include_charts)
    
    logger.info(f"✅ Excel file saved successfully: {filepath}")
    logger.info(f"   📊 Sheets: Summary, Detailed Results, Performance Stats, Module Breakdown, Score Distribution")

def _calculate_performance_statistics(results: List[RetrievalPipelineResult]) -> Dict[str, Any]:
    """Calculate comprehensive performance statistics"""
    if not results:
        return {}
    
    # Basic stats
    total_queries = len(results)
    processing_times = [r.processing_time for r in results]
    passage_counts = [len(r.top_passages) for r in results]
    
    # Extract module times
    module_times = {}
    for result in results:
        if result.statistics and 'module_times' in result.statistics:
            for module, time_val in result.statistics['module_times'].items():
                if module not in module_times:
                    module_times[module] = []
                module_times[module].append(time_val)
    
    # Calculate stats
    stats = {
        'Total Queries': total_queries,
        'Successful Queries': sum(1 for r in results if len(r.top_passages) > 0),
        'Failed Queries': sum(1 for r in results if len(r.top_passages) == 0),
        'Success Rate (%)': round((sum(1 for r in results if len(r.top_passages) > 0) / total_queries * 100), 2),
        
        'Avg Processing Time (s)': round(sum(processing_times) / len(processing_times), 2),
        'Min Processing Time (s)': round(min(processing_times), 2),
        'Max Processing Time (s)': round(max(processing_times), 2),
        'Total Processing Time (s)': round(sum(processing_times), 2),
        
        'Avg Passages per Query': round(sum(passage_counts) / len(passage_counts), 1),
        'Min Passages per Query': min(passage_counts),
        'Max Passages per Query': max(passage_counts),
        'Total Passages Retrieved': sum(passage_counts)
    }
    
    # Add module-specific stats
    for module, times in module_times.items():
        if times:
            stats[f'Avg {module.title()} Time (s)'] = round(sum(times) / len(times), 2)
            stats[f'{module.title()} % of Total'] = round((sum(times) / sum(processing_times) * 100), 1)
    
    return stats

def _create_module_breakdown_data(results: List[RetrievalPipelineResult]) -> List[Dict[str, Any]]:
    """Create data cho module breakdown analysis"""
    module_data = []
    
    for result in results:
        if result.statistics and 'module_times' in result.statistics:
            module_times = result.statistics['module_times']
            
            row = {
                'Query ID': result.query_id,
                'Total Time': round(result.processing_time, 2)
            }
            
            # Add individual module times
            for module, time_val in module_times.items():
                row[f'{module.title()} Time'] = round(time_val, 2)
                row[f'{module.title()} %'] = round((time_val / result.processing_time * 100), 1) if result.processing_time > 0 else 0
            
            module_data.append(row)
    
    return module_data

def _create_score_distribution_data(results: List[RetrievalPipelineResult]) -> List[Dict[str, Any]]:
    """Create data cho score distribution analysis"""
    score_data = []
    
    for result in results:
        for passage in result.top_passages:
            metadata = passage.get('metadata', {})
            
            score_data.append({
                'Query ID': result.query_id,
                'Passage ID': passage.get('passage_id', ''),
                'Rank': passage.get('rank', 0),
                'Final Score': passage.get('score', 0),
                'Retrieval Score': metadata.get('hybrid_retrieval_score', 0),
                'Support Score': metadata.get('support_score', 0),
                'Supporting Triples Count': metadata.get('supporting_triples_count', 0),
                'Score Difference': passage.get('score', 0) - metadata.get('hybrid_retrieval_score', 0)
            })
    
    return score_data

def _format_excel_workbook(filepath: Path, include_charts: bool = True):
    """Apply formatting to Excel workbook"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        
        wb = load_workbook(filepath)
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Header row formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add color scales for score columns
            if sheet_name == "Detailed Results":
                # Find score columns
                score_columns = []
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value and 'score' in str(cell.value).lower():
                        score_columns.append(col_idx)
                
                # Apply color scale to score columns
                for col_idx in score_columns:
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    color_rule = ColorScaleRule(
                        start_type='min', start_color='F8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                        end_type='max', end_color='63BE7B'
                    )
                    ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}', color_rule)
        
        wb.save(filepath)
        logger.info("   ✅ Excel formatting applied successfully")
        
    except Exception as e:
        logger.warning(f"⚠️ Could not apply Excel formatting: {e}")

# Add method to RetrievalPipelineOrchestrator class
def save_results_to_excel(self, results: List[RetrievalPipelineResult], 
                         filepath: Path, include_charts: bool = True):
    """
    Save pipeline results to formatted Excel file
    
    Args:
        results (List[RetrievalPipelineResult]): Results to save
        filepath (Path): Excel output path
        include_charts (bool): Include charts and advanced formatting
    """
    save_results_to_excel(results, filepath, include_charts)

# ==================== TEST FUNCTIONS ====================

def test_pipeline_orchestrator():
    """Test Pipeline Orchestrator với sample query"""
    print("🧪 TESTING PIPELINE ORCHESTRATOR")
    print("=" * 50)
    
    try:
        # Create config
        config = create_default_pipeline_config()
        config.max_final_passages = 5  # Reduce cho test
        
        # Initialize orchestrator
        orchestrator = RetrievalPipelineOrchestrator(config)
        
        # Test single query
        test_query = "Lợi ích của táo cho sức khỏe"
        result = orchestrator.process_single_query(test_query, "TEST_001")
        
        print(f"✅ Test successful!")
        print(f"   Query: {result.query}")
        print(f"   Passages found: {len(result.top_passages)}")
        print(f"   Processing time: {result.processing_time:.2f}s")
        print(f"   Method: {result.method}")
        
        # Show top passages
        print(f"\n🏆 Top passages:")
        for passage in result.top_passages[:3]:
            print(f"   - {passage['passage_id']}: {passage['score']:.3f}")
            print(f"     {passage['text'][:80]}...")
        
        # Test evaluation format
        eval_format = result.get_evaluation_format()
        print(f"\n📊 Evaluation format:")
        print(f"   Query ID: {eval_format['query_id']}")
        print(f"   Method: {eval_format['method']}")
        print(f"   Passages count: {len(eval_format['top_passages'])}")
        
        # Close orchestrator
        orchestrator.close()
        
        return result
        
    except Exception as e:
        print(f"❌ Test failed: {e}")
        logger.exception("Test error:")
        return None

if __name__ == "__main__":
    test_pipeline_orchestrator()
